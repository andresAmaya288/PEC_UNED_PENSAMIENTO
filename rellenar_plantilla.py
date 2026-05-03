#!/usr/bin/env python3
"""
Script para rellenar la plantilla Excel con los resultados del experimento.
Lee los archivos CSV (P01.csv, P02.csv, etc.) y genera un registro consolidado
en la plantilla blanco_registro_experimento_insight(1).xlsx

Uso:
    python rellenar_plantilla.py <carpeta_csv> <archivo_plantilla>

Ejemplo:
    python rellenar_plantilla.py ./Respuestas "./plantilla en blanco_registro_experimento_insight(1).xlsx"
"""

import os
import sys
import glob
import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime


def cargar_datos_csv(ruta_carpeta):
    """
    Carga todos los archivos CSV de participantes.
    
    Args:
        ruta_carpeta (str): Ruta a la carpeta con archivos CSV
        
    Returns:
        dict: Datos por participante
    """
    archivos_csv = sorted(glob.glob(os.path.join(ruta_carpeta, 'P*.csv')))
    
    if not archivos_csv:
        print(f"❌ No se encontraron archivos CSV (P01.csv, P02.csv, etc.) en '{ruta_carpeta}'")
        return None
    
    print(f"📁 Encontrados {len(archivos_csv)} archivo(s) CSV")
    
    datos_participantes = {}
    for archivo in archivos_csv:
        nombre_archivo = os.path.basename(archivo)
        codigo = nombre_archivo.replace('.csv', '').strip()
        
        df = pd.read_csv(archivo)
        datos_participantes[codigo] = df
        print(f"   ✓ {nombre_archivo} ({len(df)} registros)")
    
    return datos_participantes


def calcular_estadisticas(df_participante):
    """
    Calcula estadísticas de un participante.
    
    Args:
        df_participante (pd.DataFrame): Datos del participante
        
    Returns:
        dict: Estadísticas calculadas
    """
    stats = {
        'total_problemas': len(df_participante),
        'problemas_resueltos': (df_participante['Resuelto'] == 'S').sum(),
        'tasa_resolucion': (df_participante['Resuelto'] == 'S').sum() / len(df_participante) * 100,
        'tiempo_promedio_seg': df_participante['Tiempo_Segundos'].mean(),
        'tiempo_promedio_min': df_participante['Tiempo_Segundos'].mean() / 60,
        'tiempo_min': df_participante['Tiempo_Segundos'].min(),
        'tiempo_max': df_participante['Tiempo_Segundos'].max(),
    }
    
    # Estadísticas por bloque
    for bloque in sorted(df_participante['Bloque'].unique()):
        df_bloque = df_participante[df_participante['Bloque'] == bloque]
        resueltos_bloque = (df_bloque['Resuelto'] == 'S').sum()
        total_bloque = len(df_bloque)
        
        stats[f'bloque_{bloque}_resueltos'] = resueltos_bloque
        stats[f'bloque_{bloque}_total'] = total_bloque
        stats[f'bloque_{bloque}_tasa'] = resueltos_bloque / total_bloque * 100
        
        # Por tipo
        for tipo in sorted(df_bloque['Tipo_Problema'].unique()):
            df_tipo = df_bloque[df_bloque['Tipo_Problema'] == tipo]
            resueltos_tipo = (df_tipo['Resuelto'] == 'S').sum()
            stats[f'b{bloque}_tipo_{tipo}_resueltos'] = resueltos_tipo
            stats[f'b{bloque}_tipo_{tipo}_total'] = len(df_tipo)
            stats[f'b{bloque}_tipo_{tipo}_tiempo'] = df_tipo['Tiempo_Segundos'].mean() / 60
    
    return stats


def rellenar_excel(datos_participantes, archivo_plantilla, carpeta_salida='./'):
    """
    Rellena la plantilla Excel con los datos de los participantes.
    
    Args:
        datos_participantes (dict): Datos cargados de los CSVs
        archivo_plantilla (str): Ruta a la plantilla Excel
        carpeta_salida (str): Carpeta donde guardar el archivo relleno
    """
    if not os.path.exists(archivo_plantilla):
        print(f"❌ No se encontró la plantilla: {archivo_plantilla}")
        return
    
    print(f"\n📄 Cargando plantilla: {archivo_plantilla}")
    
    # Copiar plantilla
    nombre_salida = f"registro_resultados_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    ruta_salida = os.path.join(carpeta_salida, nombre_salida)
    
    # Leer la plantilla original
    wb = openpyxl.load_workbook(archivo_plantilla)
    ws = wb.active
    
    print(f"✓ Plantilla cargada: {ws.title}")
    
    # Preparar estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    
    # Encontrar fila de inicio para datos (generalmente después de los encabezados)
    fila_inicio = 3  # Ajusta según tu plantilla
    
    # Rellenar datos de participantes
    fila_actual = fila_inicio
    for codigo, df in sorted(datos_participantes.items()):
        stats = calcular_estadisticas(df)
        
        # Columna A: Código participante
        ws[f'A{fila_actual}'] = codigo
        
        # Columna B: Total problemas
        ws[f'B{fila_actual}'] = stats['total_problemas']
        
        # Columna C: Problemas resueltos
        ws[f'C{fila_actual}'] = stats['problemas_resueltos']
        
        # Columna D: Tasa de resolución
        ws[f'D{fila_actual}'] = round(stats['tasa_resolucion'], 1)
        
        # Columna E: Tiempo promedio (minutos)
        ws[f'E{fila_actual}'] = round(stats['tiempo_promedio_min'], 2)
        
        # Columna F-I: Bloque 1 por tipo
        if 'b1_tipo_A_resueltos' in stats:
            ws[f'F{fila_actual}'] = f"{stats['b1_tipo_A_resueltos']}/{stats['b1_tipo_A_total']}"
        if 'b1_tipo_B_resueltos' in stats:
            ws[f'G{fila_actual}'] = f"{stats['b1_tipo_B_resueltos']}/{stats['b1_tipo_B_total']}"
        if 'b1_tipo_C_resueltos' in stats:
            ws[f'H{fila_actual}'] = f"{stats['b1_tipo_C_resueltos']}/{stats['b1_tipo_C_total']}"
        if 'b1_tipo_D_resueltos' in stats:
            ws[f'I{fila_actual}'] = f"{stats['b1_tipo_D_resueltos']}/{stats['b1_tipo_D_total']}"
        
        # Columna J-M: Bloque 2 por tipo
        if 'b2_tipo_A_resueltos' in stats:
            ws[f'J{fila_actual}'] = f"{stats['b2_tipo_A_resueltos']}/{stats['b2_tipo_A_total']}"
        if 'b2_tipo_B_resueltos' in stats:
            ws[f'K{fila_actual}'] = f"{stats['b2_tipo_B_resueltos']}/{stats['b2_tipo_B_total']}"
        if 'b2_tipo_C_resueltos' in stats:
            ws[f'L{fila_actual}'] = f"{stats['b2_tipo_C_resueltos']}/{stats['b2_tipo_C_total']}"
        if 'b2_tipo_D_resueltos' in stats:
            ws[f'M{fila_actual}'] = f"{stats['b2_tipo_D_resueltos']}/{stats['b2_tipo_D_total']}"
        
        fila_actual += 1
        print(f"✓ {codigo}: {stats['problemas_resueltos']}/{stats['total_problemas']} resueltos ({stats['tasa_resolucion']:.1f}%)")
    
    # Guardar archivo
    Path(carpeta_salida).mkdir(parents=True, exist_ok=True)
    wb.save(ruta_salida)
    
    print(f"\n✓ Archivo generado: {ruta_salida}")
    return ruta_salida


def generar_resumen_texto(datos_participantes):
    """
    Genera un resumen en texto de los resultados.
    
    Args:
        datos_participantes (dict): Datos de los participantes
    """
    print("\n" + "=" * 70)
    print("📊 RESUMEN DE RESULTADOS")
    print("=" * 70)
    
    estadisticas_globales = {
        'total_participantes': len(datos_participantes),
        'total_resueltos': 0,
        'total_intentos': 0,
        'tasa_global': 0
    }
    
    for codigo, df in sorted(datos_participantes.items()):
        resueltos = (df['Resuelto'] == 'S').sum()
        total = len(df)
        tasa = resueltos / total * 100
        
        estadisticas_globales['total_resueltos'] += resueltos
        estadisticas_globales['total_intentos'] += total
        
        print(f"\n{codigo}:")
        print(f"  ├─ Resueltos: {resueltos}/{total} ({tasa:.1f}%)")
        print(f"  ├─ Bloque 1: {(df[df['Bloque']==1]['Resuelto']=='S').sum()}/4")
        print(f"  └─ Bloque 2: {(df[df['Bloque']==2]['Resuelto']=='S').sum()}/4")
    
    tasa_global = estadisticas_globales['total_resueltos'] / estadisticas_globales['total_intentos'] * 100
    
    print("\n" + "=" * 70)
    print(f"📈 ESTADÍSTICAS GLOBALES:")
    print(f"   Participantes: {estadisticas_globales['total_participantes']}")
    print(f"   Total resueltos: {estadisticas_globales['total_resueltos']}/{estadisticas_globales['total_intentos']}")
    print(f"   Tasa global: {tasa_global:.1f}%")
    print("=" * 70)


def main():
    """Función principal."""
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    
    carpeta_csv = sys.argv[1]
    archivo_plantilla = sys.argv[2] if len(sys.argv) > 2 else "plantilla en blanco_registro_experimento_insight(1).xlsx"
    
    print("=" * 70)
    print("RELLENAR PLANTILLA EXCEL - EXPERIMENTO DE CERILLAS")
    print("=" * 70)
    print()
    
    # Cargar datos CSV
    print("📊 Cargando archivos CSV...")
    datos_participantes = cargar_datos_csv(carpeta_csv)
    
    if datos_participantes is None:
        sys.exit(1)
    
    # Generar resumen
    generar_resumen_texto(datos_participantes)
    
    # Rellenar plantilla
    print("\n📝 Rellenando plantilla Excel...")
    rellenar_excel(datos_participantes, archivo_plantilla, carpeta_salida='./registro_excel')
    
    print("\n✅ PROCESO COMPLETADO")


if __name__ == '__main__':
    main()
